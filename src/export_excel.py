"""product_master / daily_transactions / weekly_forecast の3シートを
単一のExcel(.xlsx)へ書き出すエントリポイント。

python -m src.export_excel [--output PATH]
"""

import argparse
from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from src.common.config import SETTINGS, Settings
from src.daily_transactions import generate_daily_transactions
from src.product_master import generate_product_master
from src.weekly_forecast import generate_weekly_forecast

BASE_DIR = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT_PATH = BASE_DIR / "data" / "synthetic" / "inbound_demand_dummy.xlsx"

FONT_NAME = "Arial"


def _style_sheet(worksheet, df: pd.DataFrame) -> None:
    header_font = Font(name=FONT_NAME, bold=True)
    body_font = Font(name=FONT_NAME)

    for cell in worksheet[1]:
        cell.font = header_font
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font

    worksheet.freeze_panes = "A2"

    for i, col in enumerate(df.columns, start=1):
        # Series.astype(str).map(len) はpandasのバージョン・文字列dtypeバックエンド
        # (numpy object vs pyarrow-backed string)によって挙動が変わり得るため、
        # 素のPythonループでlen(str(v))を取ることでバックエンド差異を回避する。
        body_max = max((len(str(v)) for v in df[col]), default=0)
        max_len = max(len(str(col)), body_max)
        worksheet.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 24)


def build_workbook(settings: Settings = SETTINGS) -> dict:
    product_master = generate_product_master(settings)
    daily_transactions = generate_daily_transactions(settings)
    weekly_forecast = generate_weekly_forecast(daily_transactions, settings)
    return {
        "product_master": product_master,
        "daily_transactions": daily_transactions,
        "weekly_forecast": weekly_forecast,
    }


def save_workbook(sheets: dict, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            _style_sheet(writer.sheets[sheet_name], df)


def build_excel_bytes(sheets: dict) -> bytes:
    """st.download_button用にメモリ上でxlsxを組み立ててbytesを返す(save_workbookのBytesIO版)。"""
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            _style_sheet(writer.sheets[sheet_name], df)
    return buffer.getvalue()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="需給予測ダミーデータ生成")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT_PATH))
    args = parser.parse_args()

    sheets = build_workbook()
    save_workbook(sheets, Path(args.output))
    for name, df in sheets.items():
        print(f"{name}: {len(df)} rows")
    print(f"-> {args.output}")
